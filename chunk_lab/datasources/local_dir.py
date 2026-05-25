"""Load local directory files as LangChain Documents."""

from pathlib import Path
from typing import Any

from langchain_core.documents import Document

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".md", ".xlsx", ".xls", ".csv"}
MAX_FILES = 100


def _validate_dir(dir_path: str) -> Path:
    p = Path(dir_path).resolve()
    if not p.is_dir():
        raise ValueError(f"目录不存在: {dir_path}")
    return p


def _validate_path(path: str) -> Path:
    """Validate path exists (file or directory)."""
    p = Path(path).resolve()
    if not p.exists():
        raise ValueError(f"路径不存在: {path}")
    return p


def list_local_files(
    dir_path: str,
    *,
    recursive: bool = True,
    file_types: list[str] | None = None,
) -> list[dict[str, Any]]:
    root = _validate_dir(dir_path)
    allowed = set(file_types) if file_types else SUPPORTED_EXTENSIONS
    pattern = "**/*" if recursive else "*"

    files: list[dict[str, Any]] = []
    for f in sorted(root.glob(pattern)):
        if not f.is_file():
            continue
        if f.suffix.lower() not in allowed:
            continue
        if f.is_symlink():
            try:
                f.resolve().relative_to(root)
            except ValueError:
                continue
        files.append({
            "path": str(f),
            "name": f.name,
            "ext": f.suffix.lower(),
            "size": f.stat().st_size,
        })
    return files


def _load_pdf(path: Path) -> list[Document]:
    from langchain_community.document_loaders import PyPDFLoader

    loader = PyPDFLoader(str(path))
    return loader.load()


def _load_docx(path: Path) -> list[Document]:
    from langchain_community.document_loaders import Docx2txtLoader

    loader = Docx2txtLoader(str(path))
    return loader.load()


def _load_markdown(path: Path) -> list[Document]:
    text = path.read_text(encoding="utf-8", errors="replace")
    return [Document(page_content=text, metadata={"source": str(path)})]


def _load_csv(path: Path) -> list[Document]:
    from langchain_community.document_loaders import CSVLoader

    loader = CSVLoader(str(path), encoding="utf-8")
    return loader.load()


def _load_excel(path: Path) -> list[Document]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    docs: list[Document] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append("\t".join(cells))
        text = "\n".join(rows)
        if text.strip():
            docs.append(Document(
                page_content=text,
                metadata={"source": str(path), "sheet": sheet_name},
            ))
    wb.close()
    return docs


_LOADERS = {
    ".pdf": _load_pdf,
    ".docx": _load_docx,
    ".md": _load_markdown,
    ".csv": _load_csv,
    ".xlsx": _load_excel,
    ".xls": _load_excel,
}


def load_local_directory(
    dir_path: str,
    *,
    recursive: bool = True,
    file_types: list[str] | None = None,
    max_files: int = MAX_FILES,
) -> list[Document]:
    """Load all supported files from a local directory as Documents."""
    files = list_local_files(dir_path, recursive=recursive, file_types=file_types)
    if len(files) > max_files:
        raise ValueError(
            f"目录中有 {len(files)} 个支持的文件，超过上限 {max_files}。"
            f"请缩小范围或指定 file_types 过滤。"
        )

    documents: list[Document] = []
    errors: list[dict] = []

    for file_info in files:
        fpath = Path(file_info["path"])
        ext = file_info["ext"]
        loader_fn = _LOADERS.get(ext)
        if not loader_fn:
            continue
        try:
            docs = loader_fn(fpath)
            for doc in docs:
                doc.metadata.setdefault("source", str(fpath))
                doc.metadata["filename"] = fpath.name
                doc.metadata["file_type"] = ext
            documents.extend(docs)
        except Exception as e:
            errors.append({"file": str(fpath), "error": str(e)})

    if not documents and errors:
        raise ValueError(f"所有文件解析失败: {errors}")

    return documents


def load_local_corpus_text(
    path: str,
    *,
    recursive: bool = True,
    file_types: list[str] | None = None,
    max_files: int = MAX_FILES,
) -> tuple[str, list[Document]]:
    """Load file or directory and return concatenated text + original documents.

    If path points to a single file, load that file directly.
    If path points to a directory, load all supported files within it.
    """
    p = _validate_path(path)

    if p.is_file():
        ext = p.suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"不支持的文件格式: {ext}，支持: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
            )
        loader_fn = _LOADERS.get(ext)
        if not loader_fn:
            raise ValueError(f"无法加载文件: {p.name}")
        docs = loader_fn(p)
        for doc in docs:
            doc.metadata.setdefault("source", str(p))
            doc.metadata["filename"] = p.name
            doc.metadata["file_type"] = ext
    else:
        docs = load_local_directory(
            path, recursive=recursive, file_types=file_types, max_files=max_files
        )

    parts: list[str] = []
    for doc in docs:
        header = f"# {doc.metadata.get('filename', 'unknown')}"
        parts.append(f"{header}\n\n{doc.page_content}")
    text = "\n\n---\n\n".join(parts)
    return text, docs
